from __future__ import division
from openpyxl import load_workbook
import nltk
import csv


#getting corpus data from excel:
def get_excel_data_corpus(sheet):
    headers = ['str_id', 'project_id', 'tweet_text', 'label']
    excel_data_corpus = []  # list of dictionaries ##NAME OF LIST of dict
    for row_num, row in enumerate(sheet):
        if row_num is 0:  # skip the first row (don't want headers)
            continue
        row_data = {}  # open a dictionary ##NAME OF DICTIONARY OF VALUES IN EACH ROW
        for col_num, cell in enumerate(
                row):  # populating the dictionary with header names as keys and cell content as values# populating the dictionary with header names as keys and cell content as values
            if col_num > len(headers) - 1:  # ask Max what this is
                continue
            key = headers[col_num]
            value = cell.value
            row_data[key] = value

        excel_data_corpus.append(row_data)  # adding each key-value pair to the excel_data list
    return excel_data_corpus

#getting bots data from excel:
def get_excel_data_bots(sheet):
    headers = ['bot_index','str_id','source_text']
    excel_data_bots =[]
    for row_num, row in enumerate(sheet):
        if row_num is 0:  # skip the first row (don't want headers)
            continue
        row_data = {}  # open a dictionary ##NAME OF DICTIONARY OF VALUES IN EACH ROW
        for col_num, cell in enumerate(
                row):  # populating the dictionary with header names as keys and cell content as values# populating the dictionary with header names as keys and cell content as values
            if col_num > len(headers) - 1:  # ask Max what this is
                continue
            key = headers[col_num]
            value = cell.value
            row_data[key] = value
        excel_data_bots.append(row_data)  # adding each key-value pair to the excel_data list
    return excel_data_bots

# Returns bot's strids and sources:
def find_bots_strid_source(excel_data_bots):
    bot_tweets = []
    for row_data in excel_data_bots:
        if row_data != None:
            tweet = row_data['str_id'], row_data['source_text']
            #print(type(row_data['str_id']))
            bot_tweets.append(tweet)
    #print(bot_tweets)
    return bot_tweets

#finds strid, projid, and annotation label from corpus:
def find_strid_projid_annotlabel(excel_data_corpus):
    corpus_tweets = []
    for row_data in excel_data_corpus:
        if row_data != None:
            tweet =  row_data['str_id'], row_data['project_id'], row_data['label'], row_data['tweet_text']
            corpus_tweets.append(tweet)
    #print(corpus_tweets)
    return corpus_tweets

#comparison solution: https://www.quora.com/How-can-I-do-a-comparison-of-two-lists-in-Python-with-each-value
def standardize(list_of_tuples_tweets):
    result = []
    for tuple_element in list_of_tuples_tweets:
        for element in tuple_element:
            result.append(element)
    return set(result)


def shared_strid(corpus_tweets, bot_tweets):
    shared_numbers_tuple = []
    for number in standardize(corpus_tweets):
        if number in standardize(bot_tweets):
            for str_id, project_id, label, tweet_text in corpus_tweets:
                if str_id == number:
            #print(number)
                    shared_numbers_tuple.append((str_id, project_id, label, tweet_text))
            #shared_numbers.append(list)
            #shared_numbers.append(corpus_tweets[1])
    #print(shared_numbers)
    return shared_numbers_tuple


def export_bot_tweets_corpus(shared_numbers_tuple, file_name):
    with open(file_name, 'w', newline='') as f:
        writer = csv.writer(f)        
        # write headers
        writer.writerow(['str_id', 'project_id', 'label', 'tweet_text'])
        for tweet in shared_numbers_tuple:
            for elememt in tweet:
                str_id = tweet[0].encode('ascii')
                #print(type(str_id))
                project_id = tweet[1]
                label = tweet[2]
                #clean_tweet = tweet[3].encode('utf-8')
                clean_tweet = ''.join([c for c in tweet[3] if ord(c) < 128])
            writer.writerow([str_id, project_id, label, clean_tweet])


if __name__ == '__main__':
    ##I'll use these raw_input functions in the final version of the program    
    # workbook_input = raw_input('enter spreadsheet name (should be xlsx): ') # open files
    # workbook = load_workbook('./' + workbook_input + '.xlsx')
    # sheet_input1 = raw_input('enter annotated spreadsheet name (as it appears on the tab name): ')
    # sheet_input2 = raw_input('enter master list spreadsheet name (as it appears on the tab name): ')
    # sheet1 = workbook[sheet_input1]
    # sheet2 = workbook[sheet_input2]


    workbook_bots = load_workbook('./all_strid_bots.xlsx')
    workbook_corpus = load_workbook('./FINAL-CORPUS-ExpandedFull-Altogether-NoNotes.xlsx')

    sheet_bots = workbook_bots['Bots_strid']
    sheet_corpus = workbook_corpus['Corpus']

    excel_data_bots = get_excel_data_bots(sheet_bots)
    #print(excel_data_bots[1])
    excel_data_corpus = get_excel_data_corpus(sheet_corpus)
    #print(excel_data_corpus[1])


    bot_tweets = find_bots_strid_source(excel_data_bots)
    corpus_tweets = find_strid_projid_annotlabel(excel_data_corpus)
    # #print('corpus tweets: ', corpus_tweets)



    #standardized_bot = standardize(excel_data_bots)
    # #print(standardized_bot)
    #standardized_corpus = standardize(excel_data_corpus)
    # print(standardized_corpus)


    bot_strid = shared_strid(corpus_tweets, bot_tweets)
    #print(bot_strid)
    export_bot_tweets_corpus(bot_strid, 'bots_in_corpus1.csv')



    









